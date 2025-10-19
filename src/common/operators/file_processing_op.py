"""
File Processing Agent for YOPO System

A specialized agent designed for processing, analyzing, and extracting information from various file types.
This agent implements advanced file handling capabilities including text extraction, content analysis,
structured data processing, and intelligent summarization for comprehensive file understanding.
"""

import json
import os
import asyncio
from typing import Dict, Any, Optional, List, TypedDict, Annotated
import yaml
import base64
import io
from PIL import Image

from pptx.enum.shapes import MSO_SHAPE_TYPE
from pptx import Presentation
from openpyxl import load_workbook
from tabulate import tabulate
import pandas as pd
import numpy as np
from langgraph.graph import StateGraph, START, END
from langgraph.graph.message import add_messages
from langchain_core.messages import BaseMessage, HumanMessage, AIMessage, SystemMessage
import ffmpeg
import cv2
from scenedetect import open_video, SceneManager
from scenedetect.detectors import ContentDetector

from .llm.llm_wrapped_agent import create_wrapped_agent
from .llm.llm_provider import llm_instance, LLMProvider
from .utils import MarkitdownConverter
from .base_op import BaseOperator, OperatorResult
from ..utils.logging import get_logger

logger = get_logger(name="FileProcessingOperator")

class FileProcessingState(TypedDict):
    """State object for the File Processing workflow"""
    # Input
    original_query: str
    sources: List[str]
    context: Optional[Dict[str, Any]] = None

    # File processing results
    extracted_content: List[dict]
    # LLM analysis results
    analysis_result: str
    error_msg: Optional[str] = None
    
    # Messages for LLM interaction
    messages: Annotated[List[BaseMessage], add_messages]


def prepare_content_for_analysis(content: str) -> str:
    """Prepare extracted content for LLM analysis."""
    # FIXME: compression process
    max_length = 10000  # Limit content length for LLM
    
    if len(content) > max_length:
        content = content[:max_length] + f"\n\n[Content truncated for analysis - original length: {len(content)} characters]"
    
    return content

def process_image_file(source: str):
    ext = os.path.splitext(source)[-1].lower()
    with open(source, 'rb') as image_file:
        image_data = base64.b64encode(image_file.read()).decode("utf-8")
    return [{
            "type": "image",
            "source_type": "base64",
            "data": image_data,
            "mime_type": f"image/{ext.strip('.')}"
        }]

def process_video_file(source: str):
    def _normalize(img: Image.Image, target_width: int = 512) -> Image.Image:
        w, h = img.size
        return img.resize((target_width, int(target_width * h / w)), Image.Resampling.LANCZOS).convert("RGB")
    
    def _capture_screenshot(video_file: str, timestamp: float, width: int = 320) -> Image.Image:
        out, _ = (
            ffmpeg.input(video_file, ss=timestamp)
            .filter("scale", width, -1)
            .output("pipe:", vframes=1, format="image2", vcodec="png")
            .run(capture_stdout=True, capture_stderr=True)
        )
        return Image.open(io.BytesIO(out))

    def _extract_keyframes(
        video_path: str,
        frame_interval: float = 4.0,
        max_frames: int = 100,
        target_width: int = 512,
    ):
        cap = cv2.VideoCapture(video_path)
        total, fps = int(cap.get(cv2.CAP_PROP_FRAME_COUNT)), cap.get(cv2.CAP_PROP_FPS)
        duration = total / fps if fps else 0
        cap.release()

        desired = min(max(int(duration / frame_interval) or 1, 1), max_frames)

        video = open_video(video_path)
        sm = SceneManager()
        sm.add_detector(ContentDetector())
        sm.detect_scenes(video)
        scenes = sm.get_scene_list()

        frames: List[Image.Image] = []
        if scenes:
            for i in np.linspace(0, len(scenes) - 1, min(len(scenes), desired), dtype=int):
                frames.append(_capture_screenshot(video_path, scenes[i][0].get_seconds()))
        
        while len(frames) < desired and duration:
            t = len(frames) * frame_interval
            try:
                frames.append(_capture_screenshot(video_path, t))
            except Exception as e:  # short file edge‑case
                break
        
        normalized = [_normalize(f, target_width) for f in frames]

        
        content = []
        for im in normalized:
            buf = io.BytesIO()
            im.save(buf, format="JPEG", quality=90)
            image_data = base64.b64encode(buf.getvalue()).decode("utf-8")
            content.append({
                "type": "image",
                "source_type": "base64",
                "data": image_data,
                "mime_type": f"image/jpeg"
            })
    
        return content
    
    frames_content: List[dict] = _extract_keyframes(source)

    return frames_content

def process_excel_file(source: str):
    def _df_to_md(df: pd.DataFrame) -> str:
        return tabulate(df, headers="keys", tablefmt="pipe")

    def _handle_csv(source: str) -> str:
        try:
            df = pd.read_csv(source)
        except Exception as e:
            logger.error(f"CSV read failed: {e}")
            raise

        return "CSV File Processed:\n" + _df_to_md(df)

    def _handle_xlsx(source: str) -> str:
        wb = load_workbook(source, data_only=True)
        output_parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            cells_info = []

            for row in ws.iter_rows():
                for cell in row:
                    coord = f"{cell.row}{cell.column_letter}"
                    font_rgb = (
                        cell.font.color.rgb
                        if cell.font and cell.font.color and cell.font.color.rgb
                        else None
                    )
                    fill_rgb = (
                        cell.fill.fgColor.rgb
                        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb
                        else None
                    )
                    cells_info.append(
                        {
                            "index": coord,
                            "value": cell.value,
                            "font_color": font_rgb,
                            "fill_color": fill_rgb,
                        }
                    )

            # Re‑load sheet via pandas for prettier Markdown
            df = pd.read_excel(source, sheet_name=sheet_name, engine="openpyxl")

            part = (
                f"Sheet Name: {sheet_name}\n"
                f"Cell information list:\n{cells_info}\n\n"
                f"Markdown View of the content:\n{_df_to_md(df)}\n"
                + "-" * 40
            )
            output_parts.append(part)

        return "\n".join(output_parts)

    ext = os.path.splitext(source)[-1].lower()
    if ext == '.csv':
        result: str = _handle_csv(source=source)
    else:
        result: str = _handle_xlsx(source=source)
    
    content = [{
        "type": "text",
        "text": " - Attached file content: \n\n" + result
    }]
    return content

def process_normal_file(source: str):
    with open(source, 'r', encoding='utf-8') as f:
        result: str = f.read()
    
    return [{
        "type": "text",
        "text": " - Attached file content: \n\n" + result
    }]

def process_file(source: str, converter: MarkitdownConverter):
    content = []
    ext = os.path.splitext(source)[-1].lower()

    if ext in ['.png', '.jpg', '.jpeg']:
        content.extend(process_image_file(source=source))
    elif ext in ['.mov']:
        content.extend(process_video_file(source=source))
    elif ext in ['.csv', '.xls', '.xlsx']:
        content.extend(process_excel_file(source=source))
    elif ext in ['.txt', '.py']:
        content.extend(process_normal_file(source=source))
    else:
        try:
            extracted_text = converter.convert(source).text_content
        except Exception as e:
            extracted_text = f"Failed to extract content from {source}. Error: {e}"

        content.append(
            {
                "type": "text",
                "text": " - Attached file content: \n\n" + extracted_text,
            }
        )
    
    return content


# Node Functions - Independent of the operator class
async def file_processing_node(state: FileProcessingState, converter: MarkitdownConverter) -> FileProcessingState:
    """
    Node 1: File reading and processing using PIL, markitdown, etc.
    This node doesn't use LLM - just processes files directly.
    """
    logger.info("🔄 Starting file processing node...")
    sources: List[str] = state["sources"]
    logger.info(f"📁 Processing {len(sources)} source(s): {sources}")
    
    file_contents = []
    missing_files = []
    error_contents = []
    
    for i, source in enumerate(sources):
        logger.info(f"📄 Processing file {i+1}/{len(sources)}: {source}")
        if os.path.exists(source):
            try:
                processed_content = process_file(source, converter)
                file_contents.extend(processed_content)
                logger.info(f"✅ Successfully processed {source}, extracted {len(processed_content)} content item(s)")
            except Exception as e:
                logger.error(f"❌ Error processing file {source}: {str(e)}")
                error_contents.append(str(e))
        else:
            logger.error(f"❌ File not found: {source}")
            missing_files.append(source)
        
    logger.info(f"🎯 Total extracted content items: {len(file_contents)}")
    if missing_files or error_contents:
        state.update({
            "extracted_content": file_contents,
            "error_msg": "File path not found: " + ",".join(missing_files) + "\n\n Error msg: " + "\n".join(error_contents)
        })
    else:
        state.update({
            "extracted_content": file_contents,
            "error_msg": None
        })
    
    return state

async def llm_analysis_node(state: FileProcessingState, llm_provider: LLMProvider) -> FileProcessingState:
    """
    Node 2: LLM-based understanding and analysis of the processed file content.
    """
    logger.info("🤖 Starting LLM analysis node...")
    
    original_query = state["original_query"]
    extracted_content = state["extracted_content"]
    context: Optional[Dict[str, Any]] = state["context"]
    messages = state.get("messages", [])

    user_query = f"Query: {original_query}" if context is None else f"Context: {json.dumps(context, indent=2)}\n\n Query: {original_query}"
    if extracted_content:
        try:
            messages.append({
                "role": "user",
                "content": [{"type": "text", "text": user_query}] + extracted_content
            })
            
            logger.info("🚀 Sending request to LLM provider...")
            # Get LLM response
            response = llm_provider.generate(messages)
            logger.info(f"✅ Received LLM response (length: {len(response)} chars)")
            
            state.update({
                "analysis_result": response,
                "messages": messages + [AIMessage(content=response)]
            })
            logger.info("✅ LLM analysis node completed successfully")
            
        except Exception as e:
            logger.error(f"❌ Error in LLM analysis node: {str(e)}")
            error_response = f"Analysis failed: {str(e)}"
            state.update({
                "analysis_result": error_response,
                "messages": [HumanMessage(content=f"Error: {str(e)}")]
            })
    
    return state

class FileProcessingOperator(BaseOperator):
    """
    File Processing Operator for YOPO System
    
    A specialized operator designed for processing and analyzing various file types including documents,
    images, PDFs, spreadsheets, and other common file formats. This operator combines file content
    extraction with LLM-based analysis to provide intelligent insights and answers about file contents.
    
    The operator uses a two-stage workflow:
    1. File Processing: Extract and convert file content to a standardized format
    2. LLM Analysis: Analyze the extracted content using language models to answer queries
    
    Supported file types include but are not limited to:
    - Documents: PDF, DOCX, TXT, MD
    - Images: JPG, PNG, GIF, BMP, TIFF
    - Spreadsheets: XLSX, CSV
    - Presentations: PPTX
    - Audio/Video: Basic metadata extraction
    
    Key Features:
    - Multi-format file processing and content extraction
    - Intelligent content analysis using LLM capabilities
    - Structured workflow with error handling
    - Support for batch file processing
    - Context-aware query answering about file contents
    """
    
    def __init__(self, 
                 prompt_path: str,
                 **kwargs):
        # Load all prompts from YAML
        with open(prompt_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
            system_prompt = config.get('system_prompt', None)
            description = config.get('description', None)

        # Initialize base agent
        name = "FileProcessingOperator"
        super().__init__(name=name, description=description, model_name=os.getenv("MULTIMODEL_MODEL_NAME", "gpt-4.1"), system_prompt=system_prompt, **kwargs)

        self.converter: MarkitdownConverter = MarkitdownConverter()

    def _build_workflow(self):
        """Build the LangGraph workflow for file processing."""
        workflow = StateGraph(FileProcessingState)
        
        # Create wrapper functions that pass the required parameters
        async def _file_processing_wrapper(state: FileProcessingState) -> FileProcessingState:
            return await file_processing_node(state, self.converter)
        
        async def _llm_analysis_wrapper(state: FileProcessingState) -> FileProcessingState:
            return await llm_analysis_node(state, self.llm_provider)
        
        # Add nodes
        workflow.add_node("file_processor", _file_processing_wrapper)
        workflow.add_node("llm_analyzer", _llm_analysis_wrapper)
        
        # Define the workflow edges
        workflow.add_edge(START, "file_processor")
        workflow.add_edge("file_processor", "llm_analyzer")
        workflow.add_edge("llm_analyzer", END)
        
        return workflow.compile()

    async def arun(self, query: str, context: Optional[Dict[str, Any]] = None, sources: Optional[List[str]] = []) -> OperatorResult:
        """
        Run the file processing workflow.
        
        Args:
            query: User query about the file
            context: Additional context containing file information
        """
        logger.info(f"🎯 Starting file processing for query: {query}, context: {context}, sources: {sources}")
        self.worklfow = self._build_workflow()
        try:
            # Initialize state
            initial_state = FileProcessingState(
                original_query=query,
                sources=sources,
                context=context,
                extracted_content=[],
                analysis_result="",
                error_msg=None,
                messages=[]
            )
            
            # Run the workflow
            final_state = await self.worklfow.ainvoke(initial_state)
            if final_state["analysis_result"]:
                if final_state["error_msg"]:
                    base_result = final_state.get("analysis_result") + f"However, {final_state['error_msg']}"
                    # Prepare result
                    result = OperatorResult(
                        base_result=base_result
                    )
                else:
                    base_result = final_state.get("analysis_result")
                    result = OperatorResult(
                        base_result=base_result
                    )
            else:
                base_result = final_state['error_msg']
                result = OperatorResult(
                    base_result=base_result
                )
            logger.info(f"{self.name} Operator result: {base_result}")
            return result
            
        except Exception as e:
            logger.error(f"{self.name} Operator failed: {str(e)}")
            return OperatorResult(
                base_result=f"File processing failed: {str(e)}",
                metadata={"error": str(e)}
            )

        
